from django.shortcuts import render
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404, render
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth import authenticate
from datetime import date, datetime
import logging
from django.urls import reverse
from django.template import *
from django.db.models import Q

from django.contrib.auth.models import User
from alumnos.models import Alumno, Carrera, Prevision
from fichas.models import Ficha, Registro
from permisos.models import Permiso
from usuarios.models import Usuario
from areas.models import UsuarioEspecialidad, Especialidad, Area

from django.conf import settings
from io import BytesIO

from django.http import HttpResponse
from django.views.generic import View
from django.template.loader import get_template

from .utils import render_to_pdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side



# Create your views here.


#############---------FUNCION CREAR------#################

@login_required()
def crearFicha(request):

    #PERMISOS
    if request.user.groups.filter(name__in=['Mantenedor']).exists():
        return HttpResponseRedirect(reverse("home"))
    elif request.user.groups.filter(name__in=['Profesional']).exists():
        return HttpResponseRedirect(reverse("home"))
    elif request.user.groups.filter(name__in=['Asistente Social']).exists():
        return HttpResponseRedirect(reverse("home"))
    else:

        template = "crear_ficha.html"

        if request.method == 'GET':
            return render(request, template)

        if request.method == "POST":
            try:
                alumnos_sin_ficha = Alumno.objects.filter(ficha__alumno__isnull=True) #LEFT JOIN
                #print(alumnos.query)
                filtro = request.POST.get('inputSearch')
                querys = (Q(nombre__icontains=filtro) | Q(apellido_materno__icontains=filtro))
                querys |= Q(apellido_paterno__icontains=filtro)
                querys |= Q(rut__icontains=filtro)
                alumnos = alumnos_sin_ficha.filter(querys)
                return render(request, template, {'alumnos': alumnos})

            except Exception as e:
                messages.error(request,"No fue posible encontrar alumnos sin fichas clínicas. "+repr(e))
                return render(request, "home.html")


@login_required()
def confirmarCreacionFicha(request, id_alumno=None):

    #PERMISOS
    if request.user.groups.filter(name__in=['Mantenedor']).exists():
        return HttpResponseRedirect(reverse("home"))
    elif request.user.groups.filter(name__in=['Profesional']).exists():
        return HttpResponseRedirect(reverse("home"))
    elif request.user.groups.filter(name__in=['Asistente Social']).exists():
        return HttpResponseRedirect(reverse("home"))
    else:

        template = "confirmacion_ficha.html"

        if request.method == 'GET':
            alumno = Alumno.objects.get(id = id_alumno)

            try:
                if not alumno.domicilio or not alumno.prevision or not alumno.fecha_nacimiento:
                    previsiones = Prevision.objects.all().order_by('nombre_prevision')
                    return render(request, template, {'alumno': alumno, 'previsiones': previsiones})

                fichaExiste = Ficha.objects.filter(alumno = id_alumno)
                if not fichaExiste:
                    ficha = Ficha()
                    ficha.alumno = alumno
                    ficha.save()
                    del ficha
                    del alumno
                    messages.success(request, '¡Ficha creada con éxito!')
                    return HttpResponseRedirect(reverse("alumnos:verAlumno", args=[id_alumno]))

                del alumno
                messages.error(request,"Este alumno ya posee una ficha creada.")
                return HttpResponseRedirect(reverse("alumnos:verAlumno", args=[id_alumno]))

            except Exception as e:
                messages.error(request,"No fue posible crear la ficha clínica. "+repr(e))
                return render(request, "crear_ficha.html")

        if request.method == 'POST':
            try:
                alumno = Alumno.objects.get(id = id_alumno)

                fecha_nacimiento = request.POST.get('inputFechaNacimiento')
                domicilio = request.POST.get('inputDomicilio')
                representante = request.POST.get('inputRepresentante')
                prevision = request.POST.get('inputPrevision')
                nombre_social = request.POST.get('inputNombreSocial')

                if fecha_nacimiento:
                    alumno.fecha_nacimiento = fecha_nacimiento
                if domicilio:
                    alumno.domicilio = domicilio
                if representante:
                    alumno.representante_legal = representante
                if prevision:
                    alumno.prevision = prevision
                if nombre_social:
                    alumno.nombre_social = nombre_social

                alumno.save()

                ficha = Ficha()
                ficha.alumno = alumno
                ficha.save()

                del ficha
                del alumno

                messages.success(request, '¡Ficha creada con éxito!')
                return HttpResponseRedirect(reverse("alumnos:verAlumno", args=[id_alumno]))

            except Exception as e:
                messages.error(request,"No fue posible crear la ficha clínica. "+repr(e))
                return render(request, "crear_ficha.html")


#############---------FUNCION AÑADIR REGISTRO------#################

@login_required
def crearRegistro(request, id_alumno=None):

    #PERMISOS
    if request.user.groups.filter(name__in=['Mantenedor']).exists():
        return HttpResponseRedirect(reverse("home"))
    elif request.user.groups.filter(name__in=['Administrativo']).exists():
        return HttpResponseRedirect(reverse("home"))
    else:

        if request.method == 'POST':

            try:
                if not Ficha.objects.filter(alumno_id = id_alumno):
                    messages.error(request,"No se pudo registrar en ficha clínica.")
                    return HttpResponseRedirect(reverse("alumnos:verAlumno", args=[id_alumno]))

                ficha = Ficha.objects.get(alumno_id = id_alumno)
                registro = Registro()
                registro.descripcion_atencion = request.POST.get('inputObservaciones')
                registro.decision_alumno = request.POST.get('inputDecision')
                if request.POST.get("inputMotivo") == "Asistente Social":
                    registro.motivo_atencion = "Asistente Social"
                    registro.es_asistencia_social = True
                else:
                    registro.motivo_atencion = Especialidad.objects.get(id=request.POST.get('inputMotivo')).nombreEspecialidad
                    if registro.motivo_atencion == "Asistente Social":
                        registro.es_asistencia_social = True
                registro.ficha = ficha
                profesional = Usuario.objects.get(rut = request.POST.get('inputRutProfesional'))
                registro.profesional = profesional

                registro.save()

                del registro
                messages.success(request, '¡Registro ingresado exitosamente!')
                return HttpResponseRedirect(reverse("alumnos:verAlumno", args=[id_alumno]))

            except Exception as e:
                messages.error(request,"No fue posible ingresar registro. "+repr(e))
                return HttpResponseRedirect(reverse("alumnos:verAlumno", args=[id_alumno]))



#############---------FUNCION CREAR PDF FICHA CLINICA------#################

class GeneratePdf(View):
    def get(self, request, id_alumno=None):

        #PERMISOS
        if request.user.groups.filter(name__in=['Mantenedor']).exists():
            return HttpResponseRedirect(reverse("home"))
        elif request.user.groups.filter(name__in=['Profesional']).exists():
            return HttpResponseRedirect(reverse("home"))
        elif request.user.groups.filter(name__in=['Asistente Social']).exists():
            return HttpResponseRedirect(reverse("home"))
        else:

            template = get_template('pdf/ficha_pdf.html')
            ficha = Ficha.objects.get(alumno = id_alumno)
            registros = Registro.objects.filter(ficha = ficha)
            #pdf = render_to_pdf('pdf/ficha_pdf.html', {"registros": registros, "ficha": ficha})
            #return HttpResponse(pdf, content_type='application/pdf')
            context = {
                "registros": registros,
                "ficha": ficha,
            }
            html = template.render(context)
            pdf = render_to_pdf('pdf/ficha_pdf.html', context)
            if pdf:
                response = HttpResponse(pdf, content_type='application/pdf')
                filename = "Ficha_%s.pdf" %(ficha.alumno.rut)
                content = "inline; filename='%s'" %(filename)
                download = request.GET.get("download")
                if download:
                    content = "attachment; filename='%s'" %(filename)
                response['Content-Disposition'] = content
                return response
            return HttpResponse("Not found")


#############---------FUNCION CREAR REPORTES------#################

@login_required()
def reportes(request):
    #PERMISOS
    if request.user.groups.filter(name__in=['Mantenedor']).exists():
        return HttpResponseRedirect(reverse("home"))
    elif request.user.groups.filter(name__in=['Profesional']).exists():
        return HttpResponseRedirect(reverse("home"))
    elif request.user.groups.filter(name__in=['Asistente Social']).exists():
        return HttpResponseRedirect(reverse("home"))
    elif request.user.groups.filter(name__in=['Administrativo']).exists():
        return HttpResponseRedirect(reverse("home"))
    else:
        return render(request, 'reportes.html', {})




class ReporteExcel(View):

    def get(self,request,*args,**kwargs):

        #PERMISOS
        if request.user.groups.filter(name__in=['Mantenedor']).exists():
            return HttpResponseRedirect(reverse("home"))
        elif request.user.groups.filter(name__in=['Profesional']).exists():
            return HttpResponseRedirect(reverse("home"))
        elif request.user.groups.filter(name__in=['Asistente Social']).exists():
            return HttpResponseRedirect(reverse("home"))
        else:

            wb = Workbook()
            ws = wb.active

            # formato de bordes para celdas
            borde = Border(left=Side(border_style="thin", color='141414'),
                            right=Side(border_style="thin", color='141414'),
                            top=Side(border_style="thin", color='141414'),
                            bottom=Side(border_style="thin", color='141414'))

            ws['B1'] = 'Reporte de Fichas Clínicas'
            ws.merge_cells('B1:C1')
            ws['B1'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")
            #   bordes
            ws['B1'].border = borde
            ws['C1'].border = borde
            ws['B2'].border = borde
            ws['C2'].border = borde
            
            #   tamaño de columnas
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 45
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 25

            #   headers
            ws['B4'] = 'Rut'
            ws['C4'] = 'Apellido Paterno'
            ws['D4'] = 'Apellido Materno'
            ws['E4'] = 'Nombres'
            ws['F4'] = 'Carrera'
            ws['G4'] = 'N° Folio Ficha Clínica'
            ws['H4'] = 'Fecha del último registro'
            # ws.merge_cells('H:J')

            #   centrar texto de headers
            ws['B4'].alignment = Alignment(horizontal='center')
            ws['C4'].alignment = Alignment(horizontal='center')
            ws['D4'].alignment = Alignment(horizontal='center')
            ws['E4'].alignment = Alignment(horizontal='center')
            ws['F4'].alignment = Alignment(horizontal='center')
            ws['G4'].alignment = Alignment(horizontal='center')
            ws['H4'].alignment = Alignment(horizontal='center')

            #   bordes de headers
            ws['B4'].border = borde
            ws['C4'].border = borde
            ws['D4'].border = borde
            ws['E4'].border = borde
            ws['F4'].border = borde
            ws['G4'].border = borde
            ws['H4'].border = borde

            #   background de headers color gris
            ws['B4'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")
            ws['C4'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")
            ws['D4'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")
            ws['E4'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")
            ws['F4'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")
            ws['G4'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")
            ws['H4'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")

            #   obtener datos
            listaFichas = Ficha.objects.select_related('alumno')
            contador = 5    #   fila en la cual empieza a llenar datos
            for ficha in listaFichas:
                listaRegistros = Registro.objects.filter(ficha_id = ficha.pk).order_by("-fecha_creacion")
                if (len(listaRegistros) > 0):
                    registro = listaRegistros[0]    #   obtener solo el último registro realizado
                    ws.cell(row = contador, column = 8).value = registro.fecha_creacion.strftime('%d-%m-%Y')
                else:   #   en caso de que no cuente con ningun registro
                    ws.cell(row = contador, column = 8).value = " - "

                #   imprimir datos en excel
                ws.cell(row = contador, column = 2).value = ficha.alumno.rut
                ws.cell(row = contador, column = 3).value = ficha.alumno.apellido_paterno
                ws.cell(row = contador, column = 4).value = ficha.alumno.apellido_materno
                ws.cell(row = contador, column = 5).value = ficha.alumno.nombre
                ws.cell(row = contador, column = 6).value = ficha.alumno.carrera
                ws.cell(row = contador, column = 7).value = ficha.numero_folio

                #   centrar datos de las columnas
                ws.cell(row = contador, column = 2).alignment = Alignment(horizontal='center')
                ws.cell(row = contador, column = 3).alignment = Alignment(horizontal='center')
                ws.cell(row = contador, column = 4).alignment = Alignment(horizontal='center')
                ws.cell(row = contador, column = 5).alignment = Alignment(horizontal='center')
                ws.cell(row = contador, column = 6).alignment = Alignment(horizontal='center')
                ws.cell(row = contador, column = 7).alignment = Alignment(horizontal='center')
                ws.cell(row = contador, column = 8).alignment = Alignment(horizontal='center')

                #   darles bordes a las celdas
                ws.cell(row = contador, column = 2).border = borde
                ws.cell(row = contador, column = 3).border = borde
                ws.cell(row = contador, column = 4).border = borde
                ws.cell(row = contador, column = 5).border = borde
                ws.cell(row = contador, column = 6).border = borde
                ws.cell(row = contador, column = 7).border = borde
                ws.cell(row = contador, column = 8).border = borde


                contador = contador + 1


            ws['B2'] = 'Total de Fichas Clínicas existentes: ' + str(len(listaFichas)) #terminar esto
            ws.merge_cells('B2:C2')
            ws['B2'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")


            nombre_archivo = "Reporte.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "atachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response


"""
#############---------FUNCION BORRAR------#################

@login_required()
def borrarFicha(request):

#############---------FUNCION MODIFICAR------#################

@login_required()
def updateFicha(request):


#############---------FUNCION MOSTRAR------#################

@login_required()
def verFicha(request):


#############---------FUNCION LISTAR------#################

@login_required()
def listarFichas(request):


#############---------FUNCION BUSCAR------#################

@login_required()
def buscarFicha(request):
"""
